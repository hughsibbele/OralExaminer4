"""
Creates the Oral Examiner 4.0 spreadsheet template as an .xlsx file.
Tabs: Database, Config, Prompts, Questions, Logs
Run: python3 create_template.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# --- Styles ---
header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="722F37", end_color="722F37", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
body_font = Font(name="Arial", size=10)
body_align = Alignment(vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="D4A84B"),
    right=Side(style="thin", color="D4A84B"),
    top=Side(style="thin", color="D4A84B"),
    bottom=Side(style="thin", color="D4A84B"),
)

def style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}1"


# ============================================================
# Tab 1: Database
# ============================================================
ws_db = wb.active
ws_db.title = "Database"

db_headers = [
    "Timestamp",
    "Student Name",
    "Session ID",
    "Paper",
    "Status",
    "Defense Started",
    "Call Length (s)",
    "Transcript",
    "AI Multiplier",
    "AI Comment",
    "Instructor Notes",
    "Final Grade",
    "Conversation ID",
    "Selected Questions",
]
ws_db.append(db_headers)
style_header(ws_db, len(db_headers))

# Column widths
db_widths = [18, 18, 14, 22, 14, 18, 12, 22, 12, 22, 18, 12, 14, 18]
for i, w in enumerate(db_widths, 1):
    ws_db.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# Tab 2: Config
# ============================================================
ws_cfg = wb.create_sheet("Config")

ws_cfg.append(["Key", "Value"])
style_header(ws_cfg, 2)

# Non-secret config values (secrets go in Script Properties via Setup Wizard)
config_rows = [
    ["gemini_model", "gemini-3-flash-preview"],
    ["max_paper_length", "15000"],
    ["content_questions_count", "2"],
    ["process_questions_count", "1"],
    ["min_call_length", "60"],
    ["app_title", "Oral Defense Portal"],
    ["app_subtitle", ""],
    ["avatar_url", ""],
]

note_font = Font(name="Arial", size=10, italic=True, color="888888")
ws_cfg.append(["# API keys are stored securely via the Setup Wizard (not here)", ""])
ws_cfg.cell(row=2, column=1).font = note_font
ws_cfg.cell(row=2, column=2).font = note_font

for row in config_rows:
    ws_cfg.append(row)

ws_cfg.column_dimensions["A"].width = 35
ws_cfg.column_dimensions["B"].width = 45


# ============================================================
# Tab 3: Prompts  (sample Chekhov prompts — customize for your course)
# ============================================================
ws_prompts = wb.create_sheet("Prompts")

ws_prompts.append(["prompt_name", "prompt_text"])
style_header(ws_prompts, 2)

prompts = [
    [
        "grading_system_prompt",
        'You are an oral exam grading assistant. Your role is to assess transcripts of student oral defenses alongside the essays the students are defending. You produce two outputs: (1) a student-facing score for each rubric element (some scored 1-3, others 1-5), and (2) a grade multiplier the teacher can use to adjust essay grades. Your persona should be direct, clear, concise, analytical and rigorous. Always ground your assessment in specific evidence from the transcript and essay \u2014 cite particular exchanges or passages when justifying scores.',
    ],
    [
        "grading_rubric",
        """RUBRIC \u2014 Score each element on the scale indicated (some are capped at 3, others go to 5). 3 = meets expectations (defense performance broadly consistent with the quality of the essay). Most students should score 3 on most elements.

1. PAPER KNOWLEDGE (max 3): How well does the student know their own essay?
  3 \u2014 Broadly consistent with having written the essay: can summarize their argument and discuss main ideas and evidence when prompted.
  2 \u2014 Partial but uneven knowledge: can identify the general topic and some ideas but is notably vague or inaccurate about specific arguments, evidence, or sections.
  1 \u2014 Cannot meaningfully summarize their argument, gives largely vague or generic descriptions, makes claims that contradict the essay, or struggles to discuss specific evidence or ideas. Flag for review.

2. TEXT KNOWLEDGE (comparative): Compare the student's knowledge of the source texts as demonstrated in the defense versus the essay.
  5 \u2014 Dramatically more textual knowledge than the essay. Rare.
  4 \u2014 Clearly more textual knowledge than the essay.
  3 \u2014 Broadly consistent with the essay.
  2 \u2014 Noticeably less textual knowledge than the essay.
  1 \u2014 Cannot identify or discuss quotations and references from their own essay. Flag for review.

3. CONTENT UNDERSTANDING (comparative): Compare the depth of analysis and argumentation in the defense versus the essay.
  5 \u2014 Dramatically deeper than the essay. Rare.
  4 \u2014 Clearly deeper than the essay.
  3 \u2014 Broadly consistent with the essay.
  2 \u2014 Noticeably shallower than the essay.
  1 \u2014 Cannot explain or defend the arguments in the essay. Flag for review.

4. WRITING PROCESS UNDERSTANDING (max 3): Assess the student's demonstrated understanding of writing as a process.
  3 \u2014 Demonstrates understanding of writing as a process.
  2 \u2014 Vague or generic account suggesting limited reflection.
  1 \u2014 Cannot describe or reflect on a writing process. Flag for review.

SCORING:
- Average the four scores to produce the RUBRIC AVERAGE (round to 2 decimal places).
- Convert to a GRADE MULTIPLIER: multiplier = 1.00 + (average - 3) \u00d7 0.05
- Round the multiplier to 2 decimal places.

ACADEMIC INTEGRITY FLAGS:
If ANY element scores a 1, or if the average is 1.5 or below, add an INTEGRITY FLAG section.

OUTPUT FORMAT:
Respond with EXACTLY this structure:

SCORES:
Paper Knowledge: [1-3]
Text Knowledge: [1-5]
Content Understanding: [1-5]
Writing Process: [1-3]
Average: [X.XX]
Multiplier: [X.XX]

RATIONALE:
[~200 words explaining the reasoning for each element's score. Cite specific transcript exchanges.]

INTEGRITY FLAGS:
[If applicable, describe concerns. If none, write 'None.']""",
    ],
    [
        "agent_personality",
        """You are ExaminerBot, a professional and supportive oral defense examiner.

PERSONA & STYLE:
You are respectful, encouraging, and academically rigorous. Keep responses concise for audio delivery.

VOICE DELIVERY RULES (CRITICAL):
- ONE QUESTION RULE: Never ask two questions in the same turn. Ask one, wait for answer, then follow up.
- ANCHOR RULE: If asked to repeat, repeat the EXACT question. Only rephrase if they say "I don't understand."
- PATIENCE: Do not ask "Are you there?" unless silence exceeds 10 seconds.
- BREVITY: Keep responses under 3 sentences when possible.
- NO VERBAL LISTS: Ask open-ended questions, never read multiple choice options.
- NO MARKDOWN: Never say "asterisk," "bullet," or "dash" aloud.

BOUNDARIES:
- Do not hallucinate facts about texts or essays - base questions only on the provided essay
- Do not provide personal counseling
- Do not ask about texts or topics not mentioned in the student's essay""",
    ],
    [
        "agent_examination_flow",
        """EXAMINATION STRUCTURE (approx 15-20 minutes total):

PHASE 1 - GREETING (~1 minute):
Greet the student warmly, confirm their name, and set them at ease. You already have their essay - do NOT ask them to paste or share it.

PHASE 2 - SUMMARY REQUEST (~2 minutes):
Ask for the heart of their argument: "Give me the central claim of your essay, in your own words."
If vague, prompt once: "But what specifically are you arguing about [topic they mentioned]?"

PHASE 3 - CONTENT QUESTIONS (~10 minutes):
Ask the content questions provided in the QUESTIONS TO ASK section, one at a time.
For each question:
- Ask the main question
- Listen to their response
- Ask ONE personalized follow-up that references their specific essay

Follow-up types:
- EXTEND: "How might this apply to [another aspect]?"
- DEFEND: "Someone might argue the opposite - how would you respond?"
- CONNECT: "Does this connect to anything else in the text?"
- SPECIFY: "Can you point to the exact moment?"

PHASE 4 - PROCESS QUESTION (~3 minutes):
Transition: "Great thoughts! Let's talk about your writing process."
Ask the process question provided, with one follow-up.

PHASE 5 - WRAP-UP (~2 minutes):
Thank them warmly. Offer one chance for reflection: "Is there anything about your essay or this defense you'd like to add?"
Close gracefully and end the call.""",
    ],
    [
        "first_message",
        "Welcome {student_name}. Thank you for submitting your essay. I will be conducting your oral examination today. Please tell me when you are ready to begin.",
    ],
]

for row in prompts:
    ws_prompts.append(row)

ws_prompts.column_dimensions["A"].width = 28
ws_prompts.column_dimensions["B"].width = 100

# Wrap text in prompt_text column
for row in ws_prompts.iter_rows(min_row=2, min_col=2, max_col=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")


# ============================================================
# Tab 4: Questions  (sample questions — customize for your course)
# ============================================================
ws_q = wb.create_sheet("Questions")

ws_q.append(["category", "question"])
style_header(ws_q, 2)

sample_questions = [
    # Content questions (samples — replace with your own)
    ["content", "What is the central argument of your essay, and what evidence do you find most compelling?"],
    ["content", "How does the author use a specific literary technique to advance the theme you discuss?"],
    ["content", "If you could add another piece of evidence to strengthen your argument, what would it be?"],
    ["content", "What counterargument could someone make against your thesis, and how would you respond?"],
    ["content", "How does the passage you analyzed connect to the broader themes of the work?"],
    ["content", "Which part of the text did you find most challenging to interpret, and why?"],
    # Process questions (samples — replace with your own)
    ["process", "Walk me through your writing process for this essay, from initial idea to final draft."],
    ["process", "What was the most significant revision you made, and why did you make it?"],
    ["process", "If you had more time, what would you change or develop further in your essay?"],
]

for row in sample_questions:
    ws_q.append(row)

ws_q.column_dimensions["A"].width = 14
ws_q.column_dimensions["B"].width = 90

for row in ws_q.iter_rows(min_row=2, min_col=2, max_col=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")


# ============================================================
# Tab 5: Logs
# ============================================================
ws_logs = wb.create_sheet("Logs")

ws_logs.append(["Timestamp", "Source", "Message", "Data"])
style_header(ws_logs, 4)

ws_logs.column_dimensions["A"].width = 22
ws_logs.column_dimensions["B"].width = 24
ws_logs.column_dimensions["C"].width = 40
ws_logs.column_dimensions["D"].width = 60


# ============================================================
# Save
# ============================================================
output_path = "Oral_Examiner_4.0_Template.xlsx"
wb.save(output_path)
print(f"Template saved to: {output_path}")
